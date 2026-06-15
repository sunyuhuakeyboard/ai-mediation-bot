"""从《AI电话调解机器人_资料准备字段模板.xlsx》导入/更新知识库。

业务同事直接维护 Excel（节点/标签/路由/策略/话术/组件/合规/质检 8 张表），
运行本脚本 upsert 到 PG，再调用 POST /api/v1/admin/knowledge/reload 秒级生效。
仅更新 Excel 中出现的列；系统内部字段（entry_template_id / auto_chain /
confidence_min / set_slots / dynamic_kind 等）保留库内现值，不会被覆盖。

用法: python scripts/import_knowledge_xlsx.py /path/to/模板.xlsx
"""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook  # noqa: E402

# ---- 各 sheet：中文表头 -> 模型字段 ----
NODE_MAP = {"节点ID": "node_id", "节点名称": "node_name", "所属阶段": "stage",
            "节点目标": "node_goal", "进入条件": "enter_condition", "必填槽位": "required_slots",
            "允许动作": "allowed_actions", "禁止动作": "forbidden_actions",
            "默认下一节点": "default_next", "兜底节点": "fallback_node",
            "最大重试次数": "max_retry", "备注": "remark"}
LABEL_MAP = {"标签ID": "label_id", "标签类型": "label_type", "标签名称": "label_name",
             "业务分类": "category", "典型说法": "examples", "关键词示例": "keywords",
             "风险等级": "risk_level", "建议进入节点": "suggest_node",
             "是否全局优先": "global_priority", "备注": "remark"}
ROUTE_MAP = {"路由ID": "route_id", "当前节点": "current_node", "用户意图/抗性": "intent_label",
             "槽位条件": "slot_condition", "风险等级": "risk_level", "下一节点": "next_node",
             "动作类型": "action_type", "策略ID": "strategy_id", "话术模板ID": "template_id",
             "Prompt组件": "prompt_component_ids", "优先级": "priority",
             "是否全局规则": "is_global", "备注": "remark"}
STRATEGY_MAP = {"策略ID": "strategy_id", "策略名称": "strategy_name", "适用节点": "nodes",
                "适用意图/抗性": "intents", "策略目标": "goal", "策略指令": "instruction",
                "允许动作": "allowed_actions", "禁止动作": "forbidden_actions",
                "是否需要LLM": "need_llm", "风险等级": "risk_level",
                "失败回退话术ID": "fallback_template_id"}
TEMPLATE_MAP = {"模板ID": "template_id", "适用节点": "node_id", "策略ID": "strategy_id",
                "意图/抗性": "intent_label", "话术文本": "template_text",
                "变量占位符": "variables", "是否可直出": "can_direct",
                "是否需LLM改写": "need_rewrite", "合规等级": "compliance_level",
                "质检评分": "quality_score", "备注": "remark"}
COMPONENT_MAP = {"组件ID": "component_id", "组件类型": "component_type", "适用节点": "nodes",
                 "启用条件": "enable_condition", "内容模板": "content", "优先级": "priority",
                 "是否必选": "required", "备注": "remark"}
COMPLIANCE_MAP = {"规则ID": "rule_id", "规则类型": "rule_type", "触发关键词": "trigger_keywords",
                  "适用阶段": "stage", "风险等级": "risk_level", "合规要求": "requirement",
                  "触发动作": "action", "修复话术": "repair_text", "备注": "remark"}
QC_MAP = {"质检ID": "qc_id", "质检维度": "dimension", "检查点": "checkpoint",
          "规则说明": "rule_text", "严重程度": "severity", "扣分": "deduct",
          "是否需人工复核": "manual_review", "输出字段": "output_field", "备注": "remark"}

_LIST_FIELDS = {"required_slots", "keywords", "trigger_keywords", "variables",
                "prompt_component_ids"}
_BOOL_FIELDS = {"global_priority", "is_global", "can_direct", "need_rewrite",
                "required", "manual_review"}
_INT_FIELDS = {"max_retry", "priority", "quality_score", "deduct"}
# CR001/CR005 的动态语义与行为类规则标记（Excel 不维护，导入时按已知ID补齐）
_DYNAMIC_KIND = {"CR001": "PRIVACY_PRE_IDENTITY", "CR005": "THIRD_PARTY"}
_NON_INTERCEPT = {"CR006", "CR007", "CR008"}


def _split(v) -> list:
    if v in (None, "", "无"):
        return []
    return [p.strip() for p in str(v).replace("；", ";").replace("，", ",")
            .replace("、", ",").replace("+", ",").replace(";", ",").split(",") if p.strip()]


def _parse_cond(v) -> dict:
    if v in (None, "", "无"):
        return {}
    cond = {}
    for part in _split(v) or [str(v)]:
        if "=" not in part:
            continue
        k, val = (x.strip() for x in part.split("=", 1))
        low = val.lower()
        cond[k] = True if low == "true" else False if low == "false" else val
    return cond


def _coerce(field: str, value):
    if value is None:
        return None
    if field in _LIST_FIELDS:
        return _split(value)
    if field == "slot_condition":
        return _parse_cond(value)
    if field in _BOOL_FIELDS:
        return str(value).strip() in ("是", "true", "True", "1", "Y", "y")
    if field in _INT_FIELDS:
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None
    s = str(value).strip()
    return s if s and s != "无" else None


def parse_sheet(ws, mapping: dict) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    fields = [mapping.get(h) for h in header]
    out = []
    for raw in rows[1:]:
        item = {}
        for field, value in zip(fields, raw):
            if field is None:
                continue
            coerced = _coerce(field, value)
            if coerced is not None or field in _LIST_FIELDS or field == "slot_condition":
                item[field] = coerced
        if item:
            out.append(item)
    return out


def locate_sheets(wb) -> dict[str, tuple[dict, str, str]]:
    """sheet标题关键字 -> (表头映射, 模型名, 主键字段)。"""
    plan = {}
    for ws in wb.worksheets:
        t = ws.title
        if "SOP" in t or "节点" in t:
            plan[t] = (NODE_MAP, "SopNode", "node_id")
        elif "标签" in t or "意图" in t:
            plan[t] = (LABEL_MAP, "IntentLabel", "label_id")
        elif "路由" in t or "决策" in t:
            plan[t] = (ROUTE_MAP, "DecisionRoute", "route_id")
        elif "策略" in t:
            plan[t] = (STRATEGY_MAP, "Strategy", "strategy_id")
        elif "话术" in t or "模板库" in t:
            plan[t] = (TEMPLATE_MAP, "ScriptTemplate", "template_id")
        elif "Prompt" in t or "组件" in t:
            plan[t] = (COMPONENT_MAP, "PromptComponent", "component_id")
        elif "合规" in t:
            plan[t] = (COMPLIANCE_MAP, "ComplianceRule", "rule_id")
        elif "质检" in t:
            plan[t] = (QC_MAP, "QcRule", "qc_id")
    return plan


async def _validate(session) -> list[str]:
    from sqlalchemy import select

    from app.db import models as M
    from app.knowledge.validate import validate_refs

    async def rows(model):
        objs = (await session.execute(select(model))).scalars().all()
        return [{c.key: getattr(o, c.key) for c in o.__table__.columns} for o in objs]

    return validate_refs(nodes=await rows(M.SopNode), routes=await rows(M.DecisionRoute),
                         templates=await rows(M.ScriptTemplate),
                         strategies=await rows(M.Strategy),
                         components=await rows(M.PromptComponent))


async def main(path: str, strict: bool = False) -> None:
    from app.db import models as M
    from app.db.models import Base
    from app.db.postgres import init_engine, session_factory

    wb = load_workbook(path, data_only=True)
    plan = locate_sheets(wb)
    if not plan:
        print("未识别到知识表 sheet（节点/标签/路由/策略/话术/组件/合规/质检）")
        return

    engine = init_engine()
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    sf = session_factory()
    async with sf() as session:
        for title, (mapping, model_name, pk_field) in plan.items():
            model = getattr(M, model_name)
            items = parse_sheet(wb[title], mapping)
            n = 0
            for item in items:
                pk = item.get(pk_field)
                if not pk:
                    continue
                if model_name == "ComplianceRule":
                    item.setdefault("dynamic_kind", _DYNAMIC_KIND.get(pk))
                    if pk in _NON_INTERCEPT:
                        item["intercept"] = False
                obj = await session.get(model, pk)
                if obj is None:
                    obj = model(**{pk_field: pk})
                    session.add(obj)
                for k, v in item.items():
                    if k != pk_field:
                        setattr(obj, k, v)
                n += 1
            print(f"  [{title}] -> {model.__tablename__}: upsert {n} rows")
        await session.flush()
        issues = await _validate(session)
        if issues:
            print(f"\n发现 {len(issues)} 个引用完整性问题：")
            for i in issues:
                print("  - " + i)
            if strict:
                await session.rollback()
                print("--strict 模式：已回滚，未写入任何变更。请修正Excel后重试。")
                sys.exit(1)
            print("（非strict模式：仍已写入，请尽快修正后 reload）")
        await session.commit()

    print("导入完成。调用 POST /api/v1/admin/knowledge/reload 即可秒级生效（无需重启）。")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    asyncio.run(main(sys.argv[1], strict="--strict" in sys.argv))
